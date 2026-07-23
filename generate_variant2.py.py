import re
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

INPUT_XLSX = "Style_Shifting_Study_Sheet_with_step12_metrics, ostatnie pełne badanie.xlsx"
OUTPUT_XLSX = "Style_Shifting_Study_Sheet_with_step12_metrics_VARIANT2.xlsx"
OUT_SHEET = "Variant2_BP_ICP"

BP_OPEN = "Certainly — happy to help."
BP_CLOSE = "Please let me know if you’d like a shorter version."

def norm(s: str) -> str:
    return (s or "").replace("’", "'").replace("—", "—")

def wc(s: str) -> int:
    s = (s or "").strip()
    if not s:
        return 0
    return len(re.findall(r"\b\w+(?:'\w+)?\b", s))

def count_phrases(text: str, phrases: list[str]) -> int:
    t = norm(text).lower()
    return sum(t.count(norm(p).lower()) for p in phrases)

def count_if(text: str) -> int:
    return len(re.findall(r"\bif\b", norm(text).lower()))

def split_boilerplate(full_text: str):
    t = full_text or ""
    open_present = 1 if (BP_OPEN in t or ("Certainly" in t and "happy to help" in t)) else 0
    close_present = 1 if (BP_CLOSE in t) else 0

    bp_parts = []
    if open_present:
        bp_parts.append(BP_OPEN)
    if close_present:
        bp_parts.append(BP_CLOSE)
    bp_text = "\n".join(bp_parts)

    body = t.replace(BP_OPEN, "").replace(BP_CLOSE, "")
    body = re.sub(r"\n{3,}", "\n\n", body).strip()
    return open_present, close_present, bp_text, body

ICP_REQUEST = ["could you", "would you mind", "would you", "would it be possible",
               "i was wondering if", "i'd be grateful if", "i’d be grateful if"]
ICP_GRAT = ["thank you", "thanks for", "i appreciate", "i'd appreciate", "i’d appreciate", "much appreciated"]
ICP_APO = ["i'm sorry", "i am sorry", "apologies for", "i apologize", "i apologise", "sorry about"]
ICP_EMP = ["i understand", "that makes sense", "i can see why", "i'm sorry to hear", "i’m sorry to hear"]

EHI = ["might", "may", "could", "likely", "unlikely", "it depends",
       "i'm not sure", "i am not sure", "it's possible", "it is possible",
       "can't confirm", "cannot confirm", "seems", "appears"]

COI_OTHER = ["unless", "otherwise", "alternatively", "one option is",
             "another option is", "option a", "option b", "you could also"]

wb = load_workbook(INPUT_XLSX)

required = ["response_id", "tone", "certainty", "task_type", "response_text"]
src_ws = None
hdr = None

for ws in wb.worksheets:
    headers = {str(ws.cell(1, c).value).strip().lower(): c
               for c in range(1, ws.max_column + 1)
               if ws.cell(1, c).value is not None}
    if all(k in headers for k in required):
        src_ws = ws
        hdr = headers
        break

if src_ws is None:
    raise ValueError("Nie znalazłem arkusza z nagłówkami: response_id, tone, certainty, task_type, response_text (wiersz 1).")

if OUT_SHEET in wb.sheetnames:
    del wb[OUT_SHEET]
out_ws = wb.create_sheet(OUT_SHEET)

out_headers = [
    "response_id","tone","certainty","task_type",
    "BP_open_present","BP_close_present","BP_count","BP_text","BP_words","BP_politeness_markers",
    "Body_text","Body_words",
    "ICP_request_indirectness","ICP_gratitude","ICP_apology","ICP_empathy_validation",
    "ICP_total","ICP_density_per1000",
    "EHI_epistemic","EHI_density_per1000",
    "COI_conditionality_options","COI_density_per1000"
]
out_ws.append(out_headers)

c_id, c_tone, c_cert, c_task, c_text = (hdr["response_id"], hdr["tone"], hdr["certainty"], hdr["task_type"], hdr["response_text"])

for r in range(2, src_ws.max_row + 1):
    resp = src_ws.cell(r, c_text).value
    if resp is None or str(resp).strip() == "":
        continue
    resp = str(resp)

    response_id = src_ws.cell(r, c_id).value
    tone = src_ws.cell(r, c_tone).value
    certainty = src_ws.cell(r, c_cert).value
    task_type = src_ws.cell(r, c_task).value

    open_p, close_p, bp_text, body_text = split_boilerplate(resp)

    bp_words = wc(bp_text)
    bp_pol = count_phrases(bp_text, ["certainly", "happy", "please"])

    body_words = wc(body_text)

    icp_req = count_phrases(body_text, ICP_REQUEST)
    icp_gra = count_phrases(body_text, ICP_GRAT)
    icp_apo = count_phrases(body_text, ICP_APO)
    icp_emp = count_phrases(body_text, ICP_EMP)
    icp_total = icp_req + icp_gra + icp_apo + icp_emp
    icp_density = (icp_total / body_words * 1000) if body_words else 0

    ehi = count_phrases(body_text, EHI)
    ehi_density = (ehi / body_words * 1000) if body_words else 0

    coi = count_if(body_text) + count_phrases(body_text, COI_OTHER)
    coi_density = (coi / body_words * 1000) if body_words else 0

    out_ws.append([
        response_id, tone, certainty, task_type,
        open_p, close_p, open_p + close_p, bp_text, bp_words, bp_pol,
        body_text, body_words,
        icp_req, icp_gra, icp_apo, icp_emp,
        icp_total, icp_density,
        ehi, ehi_density,
        coi, coi_density
    ])

for i, h in enumerate(out_headers, start=1):
    col = get_column_letter(i)
    out_ws.column_dimensions[col].width = 70 if h in ("BP_text", "Body_text") else 22

wb.save(OUTPUT_XLSX)
print(f"✅ Saved: {OUTPUT_XLSX}")